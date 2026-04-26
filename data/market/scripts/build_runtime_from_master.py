from __future__ import annotations

import json
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
MASTER_PATH = ROOT / "data" / "market" / "master" / "extract_frisia_data_master.xlsx"
RUNTIME_PATH = ROOT / "data" / "market" / "runtime" / "leadgen_market_data.json"


NUMERIC_FIELDS = {
    "verkaeufe_anzahl",
    "min_preis_eur_m2",
    "quantil_01_preis_eur_m2",
    "median_preis_eur_m2",
    "durchschnitt_preis_eur_m2",
    "quantil_09_preis_eur_m2",
    "max_preis_eur_m2",
    "delta_vorjahr_median_prozent",
    "efh_median_preis_eur",
    "tage_am_markt",
    "tage_am_markt_vorjahr",
    "ankaufspreis_best_deal_eur_m2",
    "median_2016",
    "median_2017",
    "median_2018",
    "median_2019",
    "median_2020",
    "median_2021",
    "median_2022",
    "median_2023",
    "median_2024",
    "median_2025",
}


def normalize_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value == "":
        return None
    return value


def normalize_record(record: dict[str, object]) -> dict[str, object]:
    output: dict[str, object] = {}

    for key, raw in record.items():
        value = normalize_value(raw)

        if key in NUMERIC_FIELDS and value is not None:
            try:
                number = float(value)
            except (TypeError, ValueError):
                output[key] = None
                continue

            if number.is_integer():
                output[key] = int(number)
            else:
                output[key] = round(number, 2)
            continue

        if key in {"leadgen_geeignet", "landingpage_geeignet"}:
            output[key] = str(value).strip().lower() == "ja" if value is not None else False
            continue

        output[key] = value

    output["postal_codes"] = [
        part.strip()
        for part in str(output.get("plz_bereiche") or output.get("plz") or "").split(",")
        if part and part.strip()
    ]
    return output


def main():
    wb = load_workbook(MASTER_PATH, read_only=True, data_only=True)
    ws = wb["imv_marktdaten"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(item) for item in next(rows)]

    records = [normalize_record(dict(zip(headers, row))) for row in rows]

    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFile": MASTER_PATH.name,
        "recordCount": len(records),
        "records": records,
    }

    RUNTIME_PATH.parent.mkdir(parents=True, exist_ok=True)
    RUNTIME_PATH.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    print(f"Runtime-Datei geschrieben: {RUNTIME_PATH}")
    print(f"Datensätze: {len(records)}")


if __name__ == "__main__":
    main()
